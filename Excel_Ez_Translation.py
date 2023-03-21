import os
from os import system
import googletrans
from googletrans import *
from openpyxl import workbook, load_workbook

global language
language = "portuguese"

def menu():
    global language
    os.system('cls')
    print("""
    Excel translator

    MENU    
    ------------------------------                                                           
    1   Translate
    2   Change language
    3   Help
    0   Exit
    ------------------------------
    """)

    try:
        var = int(input("\tOption : "))
        if var < 0  or var > 3:
            raise Exception()
    except:
        print("\n Invalid choice")
        os.system('pause')
        menu()

    if var == 1:
        translate()
        os.system('pause')
        menu()

    elif var == 2:
        print(f'\tCurrent language : {language}')
        print('\tExamples: spanish,korean,japanese...')
        language = input("\tLanguage : ")
        menu()

    elif var == 3:
        os.system('cls')
        print("""
    Excel translator

    Help
    --------------------------------------------------------------           
    Setting up the data
        - Open the "File.xlsx"
        - Place the content that you want to translate in the row A
    
    Changing the language (Default: portuguese)
        - Choose the second option to check and change the current language 
    --------------------------------------------------------------
        """)
        os.system('pause')
        menu()

    elif var == 0 :
        exit()

def translate():
    os.system('cls')
    try:
        wb = load_workbook('file.xlsx')
        ws = wb.active
        ws['B1'].value = language
        col = 1
        for i in range(ws.max_row-1):
            col += 1
            ws[f'B{col}'].value = googletrans.Translator().translate(ws[f'A{col}'].value, dest=language).text

        wb.save('file.xlsx')
        print(f'\t{col - 1} cells successfully translated to {language}')

    except:
        print("\tAn error occurred")

menu()